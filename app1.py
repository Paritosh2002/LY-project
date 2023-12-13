from tkinter import *
from PIL import Image, ImageTk
from tkinter import ttk
from tkinter import filedialog
from tkinter.filedialog import askopenfile
import os
import sys
import pdfplumber
import re
import pandas
from openpyxl import load_workbook

root = Tk()
root.title('PDF Extract')
root.geometry("1300x790")
root.resizable(False, False)

def center_screen():
    """ gets the coordinates of the center of the screen """
    global screen_height, screen_width, x_cordinate, y_cordinate
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_cordinate = int((screen_width/2) - (1300/2))
    y_cordinate = int((screen_height/2) - (790/2)-40)
    root.geometry("{}x{}+{}+{}".format(1300, 790, x_cordinate, y_cordinate))
center_screen()

canvas = Canvas(root, width=1300, height=900)
canvas.grid(columnspan=13, rowspan=15)
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(os.path.dirname(__file__))

image1_path = os.path.join(base_path, "images", "logoneww.png")
image2_path = os.path.join(base_path, "images", "bg3.jpg")
background_image = Image.open(image2_path)  
background_photo = ImageTk.PhotoImage(background_image)
canvas.create_image(0, 0, image=background_photo, anchor=NW)

frame = Frame(root)

photo = Image.open(image1_path)
photo = ImageTk.PhotoImage(photo)
canvas.create_image(650,100, image=photo)

label_excel = Label(root, text="Upload excel sheet for data entry:",height=2, width=32, font='Tahoma', bg='light blue')
label_excel.grid(row=6 , column=6, sticky='n')

def open_excel_file():
    global file_path_ex
    file_path_ex = filedialog.askopenfilename(filetypes=[("Excel Files", ["*.xlsx", "*.xls"])])
    if file_path_ex:
        label_text = "Uploaded excel file: " + file_path_ex
        label_excel.config(text=label_text)
        label_width = len(label_text) + 2  
        label_excel.config(width=label_width)

browse_text = StringVar()
button_excel = Button(root, textvariable=browse_text, command=open_excel_file, bg='light sea green', font="Tahoma", fg="black", height=2, width=35, compound="center")
browse_text.set("Select Excel file")
button_excel.grid(row=7, column=6, sticky='n')

label = Label(root, text="Upload bill with PDF extension in one of the below given formats:",height=2, width=70, font='Tahoma', bg='light blue')
label.grid(row=8 , column=6, sticky='n')

def exit_tk():
    root.destroy() 

label2 = Label(root, text="",height=2, width=50, font='Tahoma', fg='black', bg='CadetBlue1')
label2.grid(row=11, column=6, sticky='n', pady=(15, 0))
        
def extract_info_from_format_1(file):
    global file_path
    file_path1 = file_path
    with pdfplumber.open(file) as pdf:
        first_page = pdf.pages[0]
        layout_text1 = first_page.extract_text(x_tolerance=3, y_tolerance=3, layout=False, x_density=7.25, y_density=13)
    layout_text = first_page.extract_text(x_tolerance=3, y_tolerance=3, layout=True, x_density=7.25, y_density=13)
    panda1 = pandas.DataFrame(first_page.extract_table(table_settings={}))
    column_names2 = panda1.loc[(panda1[0].str.startswith('Sr', na=False)) | (panda1[0].isin(["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]))]
    
    column_names2_clean=column_names2.dropna(axis=1,how='all')
    dic = {}
    headers = column_names2_clean.iloc[0]
    new_df  = pandas.DataFrame(column_names2_clean.values[1:], columns=headers)
    header_list = headers.values.flatten().tolist()
    
    row_list1 = new_df.loc[0, :].values.flatten().tolist()
    for i in range(len(new_df.index)):
        item_list = []
        row_array = new_df.loc[i, :].values.flatten().tolist()
        for j in range(len(header_list)):
            item_split = row_array[j].split('\n', 1)[0]
            item_list.append(str(header_list[j]) + ": " + item_split)
        dic[i + 1] = item_list
    invoice_no_pattern = r"INVOICE NO\. :(\d+)"
    invoice_no_match = re.search(invoice_no_pattern, layout_text)

    if invoice_no_match:
        invoice_no = invoice_no_match.group(1)
        print("Matched Invoice Number:", invoice_no)
    else:
        print("Invoice number not found.")
    buyer_pattern = r"M/s\.(.+?) INVOICE NO"
    buyer_match = re.search(buyer_pattern, layout_text)
    if buyer_match:
        buyer_name = buyer_match.group(1).strip()
    else:
        buyer_name = "Buyer name not found"
        
    lines = layout_text1.split('\n')

    seller_name = ""

    for i, line in enumerate(lines):
        line_lower = line.lower()
        if "tax invoice" in line_lower and i < len(lines) - 1:
            seller_name = lines[i + 1].strip()
            break
        elif "bill of supply" in line_lower and i < len(lines) - 1:
            seller_name = lines[i + 1].strip()
            break
    gst_id = None
    fssai = None
    unique_gst = set()
    invoice_dates=[]
    final_amount=[]
    gst_matches = re.findall(r"\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}", layout_text)
    unique_gst.update(gst_matches)
    fssai_matches = re.findall(r"\b[0-9]{14}\b", layout_text)
    unique_fssai = fssai_matches
    date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'
    final_date = set()
    dates = re.findall(date_pattern, layout_text)

    for i in dates:
        if dates.count(i) > 1:
            final_date.add(i)
    pattern_total = r"INVOICE TOTAL (?:₹ )?(\d+\.\d{2})"

    match = re.search(pattern_total, layout_text)
    if match:
        invoice_total = match.group(1)
    else:
        invoice_total = "Invoice total not found."
        
    date_item = str(final_date)
    cleaned_date = date_item.strip("{'")[:-2].replace('/','-')
    new = file.split("/")
    file_name = new[-1]
    gst_list = list(unique_gst)
    gst_str = " , ".join(gst_list)
    fssai_list = list(unique_fssai)
    fssai_str = " , ".join(fssai_list)
    dic1_keys = list(dic)
    workbook = load_workbook(file_path_ex)
    sheet = workbook.active
    a = sheet.max_row+1
    
    for i in range(len(dic1_keys)):
        
        
        sheet.cell(column=1, row=a, value=file_name)
        sheet.cell(column=2, row=a, value=invoice_no)
        sheet.cell(column=3, row=a, value=buyer_name)
        sheet.cell(column=4, row=a, value=seller_name)
        sheet.cell(column=5, row=a, value=gst_str)
        sheet.cell(column=6, row=a, value=fssai_str)
        sheet.cell(column=7, row=a, value=str(cleaned_date))
        sheet.cell(column=8, row=a, value=float(invoice_total))
        dic_item = dic.get(dic1_keys[i])
       
        item_obj = dic_item[1].split(":")
        
        item_obj_final = item_obj[-1]
        item_obj_final_strip = item_obj_final.strip()
        
        sheet.cell(column=9, row=a, value=str(item_obj_final_strip))
        hsn_obj = dic_item[2].split(":")
        hsn_obj_final = hsn_obj[-1]
        hsn_obj_final_strip = hsn_obj_final.strip()
        sheet.cell(column=10, row=a, value=str(hsn_obj_final_strip))
       
        weight_obj = dic_item[5].split(":")
        weight_obj_final = weight_obj[-1]
        weight_obj_final_strip = weight_obj_final.strip()
        sheet.cell(column=11, row=a, value=str(weight_obj_final_strip))
        
        rate_obj = dic_item[6].split(":")
        rate_obj_final = rate_obj[-1]
        rate_obj_final_strip = rate_obj_final.strip()
        sheet.cell(column=12, row=a, value=str(rate_obj_final_strip))
        
        amt_obj = dic_item[8].split(":")
        amt_obj_final = amt_obj[-1]
        amt_obj_final_strip = amt_obj_final.strip()
        sheet.cell(column=13, row=a, value=str(amt_obj_final_strip))
        
        if column_names2_clean[6][4]=="Qty":
            qty = column_names2_clean[6][5]
        else:
            qty = column_names2_clean[5][4]
        sheet.cell(column=14, row=a, value=qty)
        workbook.save(file_path_ex)
    """a = sheet.max_row+1
    sheet.cell(column=1, row=a, value=file_name)
    sheet.cell(column=2, row=a, value=buyer_name)
    sheet.cell(column=3, row=a, value=seller_name)
    sheet.cell(column=4, row=a, value=gst_str)
    sheet.cell(column=5, row=a, value=fssai_str)
    sheet.cell(column=6, row=a, value=str(cleaned_date))
    sheet.cell(column=7, row=a, value=float(invoice_total))
    sheet.cell(column=8, row=a, value=str(dic))"""
    
    if (
        sheet.cell(column=2, row=a).value == invoice_no and 
        sheet.cell(column=1, row=a).value == file_name and
        sheet.cell(column=3, row=a).value == buyer_name and
        sheet.cell(column=4, row=a).value == seller_name and
        sheet.cell(column=5, row=a).value == gst_str and
        sheet.cell(column=6, row=a).value == fssai_str and
        sheet.cell(column=7, row=a).value == str(cleaned_date) and
        sheet.cell(column=8, row=a).value == float(invoice_total) and
        sheet.cell(column=9, row=a, value=str(item_obj_final_strip)) and 
        sheet.cell(column=10, row=a, value=str(hsn_obj_final_strip)) and 
        sheet.cell(column=11, row=a, value=str(weight_obj_final_strip)) and
        sheet.cell(column=12, row=a, value=str(rate_obj_final_strip)) and
        sheet.cell(column=13, row=a, value=str(amt_obj_final_strip)) and
        sheet.cell(column=13, row=a, value=qty)

    ):
        label2.config(text="Cosoha Format data uploaded in selected excel sheet.")
        root.after(2500, lambda: label2.config(text=""))
    else:
        label2.config(text="Cosoha Format data not uploaded")

def extract_info_from_format_2(file):
    global file_path
    file_path2 = file_path
    with pdfplumber.open(file) as pdf:
        first_page = pdf.pages[0]
        layout_text1 = first_page.extract_text(x_tolerance=3, y_tolerance=3, layout=False, x_density=7.25, y_density=13)
    layout_text = first_page.extract_text(x_tolerance=3, y_tolerance=3, layout=True, x_density=7.25, y_density=13)
    panda1 = pandas.DataFrame(first_page.extract_table(table_settings={}))
    column_names2 = panda1.loc[(panda1[0].str.startswith('Sr', na=False)) | (panda1[0].isin(["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]))]
    column_names2_clean=column_names2.dropna(axis=1,how='all')
    dic = {}
    headers = column_names2_clean.iloc[0]
    new_df  = pandas.DataFrame(column_names2_clean.values[1:], columns=headers)
    def starts_with_letter(s):
        if len(s) > 0:
            return s[0].isalpha()
        return False

    new_df = new_df[new_df['Article Description'].apply(starts_with_letter)]
    header_list = headers.values.flatten().tolist()
    row_list1 = new_df.loc[0, :].values.flatten().tolist()
    for i in range(len(new_df.index)):
        item_list = []
        row_array = new_df.loc[i, :].values.flatten().tolist()
        for j in range(len(header_list)):
            item_split = row_array[j].split('\n', 1)[0]
            item_list.append(str(header_list[j]) + ": " + item_split)
        dic[i + 1] = item_list
    
    lines = layout_text1.split('\n')
    seller_name = ""
    invoice_no_pattern = r'Inv No : \s*(.*)'
    invoice_no_match = re.search(invoice_no_pattern, layout_text)
    invoice_no=0
    if invoice_no_match:
        invoice_no = invoice_no_match.group(1)
        print("Matched Invoice Number:", invoice_no)
    else:
        print("Invoice number not found.")
    for i, line in enumerate(lines):
        if "II JAI MAHAKALI II" in line and i < len(lines) - 1:
            seller_name = lines[i + 1].strip()
            break
    lines = layout_text1.split('\n')

    buyer_name = ""
    for i, line in enumerate(lines):
        if "Shipped To" in line and i < len(lines) - 1:
            buyer_name = lines[i + 1].strip()
            break

    parts = buyer_name.split('AVENUE')
    if len(parts) > 2:
        buyer_name = 'AVENUE' + parts[1]  

    gst_id = None
    fssai = None
    unique_gst = set()
    unique_fssai = set()
    invoice_dates=[]
    final_amount=[]
    gst_matches = re.findall(r"\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}", layout_text)
    unique_gst.update(gst_matches)
    fssai_matches = re.findall(r"\b[0-9]{14}\b", layout_text)
    unique_fssai.update(fssai_matches)

    pattern = r'\b(\d{2}-\d{2}-\d{2})\b'
    match = re.search(pattern, layout_text)
    if match:
        date = match.group(1)
    else:
        date = "Date not found."

    pattern = r'Total.:\s*(.*)'
    match = re.search(pattern, layout_text)
    if match:
        extracted_string = match.group(1)
        extracted_string = extracted_string.replace(' ', '')
    else:
        extracted_string = "Invoice Total not found."

    new = file.split("/")
    file_name = new[-1]
    gst_list = list(unique_gst)
    gst_str = " , ".join(gst_list)
    fssai_list = list(unique_fssai)
    fssai_str = " , ".join(fssai_list)
    dic1_keys = list(dic)
    workbook = load_workbook(file_path_ex)
    sheet = workbook.active
    a = sheet.max_row+1
    for i in range(len(dic1_keys)):
        
        sheet.cell(column=1, row=a, value=file_name)
        sheet.cell(column=2, row=a, value=invoice_no)
        sheet.cell(column=3, row=a, value=buyer_name)
        sheet.cell(column=4, row=a, value=seller_name)
        sheet.cell(column=5, row=a, value=gst_str)
        sheet.cell(column=6, row=a, value=fssai_str)
        sheet.cell(column=7, row=a, value=str(date))
        sheet.cell(column=8, row=a, value=extracted_string)
        dic_item = dic.get(dic1_keys[i])
    
        item_obj = dic_item[1].split(":")
        
        item_obj_final = item_obj[-1]
        item_obj_final_strip = item_obj_final.strip()
       
        sheet.cell(column=9, row=a, value=str(item_obj_final_strip))
        hsn_obj = dic_item[3].split(":")
        hsn_obj_final = hsn_obj[-1]
        hsn_obj_final_strip = hsn_obj_final.strip()
        sheet.cell(column=10, row=a, value=str(hsn_obj_final_strip))
        
        weight_obj = dic_item[6].split(":")
        weight_obj_final = weight_obj[-1]
        weight_obj_final_strip = weight_obj_final.strip()
        sheet.cell(column=11, row=a, value=str(weight_obj_final_strip))
       
        rate_obj = dic_item[7].split(":")
        rate_obj_final = rate_obj[-1]
        rate_obj_final_strip = rate_obj_final.strip()
        sheet.cell(column=12, row=a, value=str(rate_obj_final_strip))
        
        amt_obj = dic_item[10].split(":")
        amt_obj_final = amt_obj[-1]
        amt_obj_final_strip = amt_obj_final.strip()
        sheet.cell(column=13, row=a, value=str(amt_obj_final_strip))
        
        qty = column_names2_clean[4][9+i]
        
        sheet.cell(column=14, row=a, value=qty)
        workbook.save(file_path_ex)
    """a = sheet.max_row+1
    sheet.cell(column=1, row=a, value=file_name)
    sheet.cell(column=2, row=a, value=buyer_name)
    sheet.cell(column=3, row=a, value=seller_name)
    sheet.cell(column=4, row=a, value=gst_str)
    sheet.cell(column=5, row=a, value=fssai_str)
    sheet.cell(column=6, row=a, value=str(date))
    sheet.cell(column=7, row=a, value=extracted_string)
    sheet.cell(column=8, row=a, value=str(dic))"""
    
    if (
        sheet.cell(column=1, row=a).value == file_name and
        sheet.cell(column=2, row=a).value == invoice_no and
        sheet.cell(column=3, row=a).value == buyer_name and
        sheet.cell(column=4, row=a).value == seller_name and
        sheet.cell(column=5, row=a).value == gst_str and
        sheet.cell(column=6, row=a).value == fssai_str and
        sheet.cell(column=7, row=a).value == str(date) and
        sheet.cell(column=8, row=a).value == extracted_string and
        sheet.cell(column=9, row=a, value=str(item_obj_final_strip)) and 
        sheet.cell(column=10, row=a, value=str(hsn_obj_final_strip)) and 
        sheet.cell(column=11, row=a, value=str(weight_obj_final_strip)) and
        sheet.cell(column=12, row=a, value=str(rate_obj_final_strip)) and
        sheet.cell(column=13, row=a, value=str(amt_obj_final_strip)) and
        sheet.cell(column=14, row=a, value=qty) 
        
    ):
        label2.config(text="Rajeshwari Format data uploaded in selected excel sheet.")
        root.after(2500, lambda: label2.config(text=""))
    else:
        label2.config(text="Rajeshwari Format data not uploaded")   
         
def extract_info_from_format_3(file):
    global file_path
    file_path3 = file_path
    with pdfplumber.open(file) as pdf:
        first_page = pdf.pages[0]
        layout_text1 = first_page.extract_text(x_tolerance=3, y_tolerance=3, layout=False, x_density=7.25, y_density=13)
    layout_text = first_page.extract_text(x_tolerance=3, y_tolerance=3, layout=True, x_density=7.25, y_density=13)
    panda1 = pandas.DataFrame(first_page.extract_table(table_settings={}))
    
    column_names2 = pandas.DataFrame()

    #column_names2 = panda1.loc[(panda1[0].str.startswith('S.N.', na=False)) | (panda1[0].isin(["1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "10."]))]
    s = {"1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "10."}
    for y,x in enumerate(panda1[0]):
            
            if x =="S.N." or x[:2] in s:
                    column_names2 = column_names2.append(panda1.iloc[y])
    column_names2_clean=column_names2.dropna(axis=1,how='all')
    
    newline_count = column_names2.iloc[1, 0].count('\n')
    items = newline_count+1
    arr=[]
    final_list=[]
    if newline_count !=0:
            first_col = column_names2.iloc[1, 1].count('\n')
            first_col_list= column_names2_clean.iloc[1, 1].split()
            number_string = ' '.join(map(str, first_col_list))

        # Calculate the length of each part
            part_length = len(number_string) // items

        # Split the string into n parts
            split_list = [number_string[i * part_length: (i + 1) * part_length] for i in range(items)]
            


            for c in range(8):
                    if c!=1:
                            
                            li = column_names2_clean.iloc[1, c].split()
                            arr.append(li)
                    else:
                            arr.append(split_list)
            temp = []
            
            for v in range(len(arr[0])):
                    temp=[]
                    for j in range(len(arr)):
                            temp.append(arr[j][v])
                    final_list.append(temp)
            

            column_names2_clean = column_names2_clean.iloc[[0]]
            new_rows_df = pandas.DataFrame(final_list, columns=column_names2_clean.columns)
            column_names2_clean = column_names2_clean.append(new_rows_df, ignore_index=True)
    
    invoice_no_pattern = r"Invoice No\. : (\d+)"
    invoice_no_match = re.search(invoice_no_pattern, layout_text)
    invoice_no=0
    if invoice_no_match:
        invoice_no = invoice_no_match.group(1)
        print("Matched Invoice Number:", invoice_no)
    else:
        print("Invoice number not found.")
    dic = {}
    headers = column_names2_clean.iloc[0]
    new_df  = pandas.DataFrame(column_names2_clean.values[1:], columns=headers)
    header_list = headers.values.flatten().tolist()
    row_list1 = new_df.loc[0, :].values.flatten().tolist()

    for i in range(len(new_df.index)):
        item_list = []
        row_array = new_df.loc[i, :].values.flatten().tolist()
        for j in range(len(header_list)):
            item_split = row_array[j].split('\n', 1)[0]
            item_list.append(str(header_list[j]) + ": " + item_split)
        dic[i + 1] = item_list
    lines = layout_text1.split('\n')
    
    seller_name = ""
    for i, line in enumerate(lines):
        if "TAX INVOICE" in line and i < len(lines) - 1:
            seller_name = lines[i + 1].strip()
            break
        elif "BILL OF SUPPLY" in line and i < len(lines) - 1:
            seller_name = lines[i + 1].strip()
            break
            
    lines = layout_text1.split('\n')
    buyer_name = ""
    for i, line in enumerate(lines):
        if "Shipped to :" in line and i < len(lines) - 1:
            buyer_name = lines[i + 1].strip()
            break
    parts = buyer_name.split('AVENUE')
    if len(parts) > 2:
        buyer_name = 'AVENUE' + parts[1]  
    gst_id = None
    fssai = None
    unique_gst = set()
    unique_fssai = set()
    invoice_dates=[]
    final_amount=[]
    gst_matches = re.findall(r"\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}", layout_text)
    unique_gst.update(gst_matches)
    fssai_matches = re.findall(r"\b[0-9]{14}\b", layout_text)
    unique_fssai.update(fssai_matches)
    pattern = r'\b(\d{2}-\d{2}-\d{4})\b'

    match = re.search(pattern, layout_text)
    if match:
        date = match.group(1)
    else:
        date = "Date not found in the text."

    pattern = r'Grand Total\s*(.*)'
    match = re.search(pattern, layout_text)

    if match:
        extracted_string = match.group(1)
        parts2=extracted_string.split(' ')
        filtered_list = [value for value in parts2 if value != '']
        invoice_total=filtered_list[-1]
    else:
        invoice_total = "Invoice Total not found in the text."

    new = file.split("/")
    file_name = new[-1]
    gst_list = list(unique_gst)
    gst_str = " , ".join(gst_list)
    fssai_list = list(unique_fssai)
    fssai_str = " , ".join(fssai_list)
    dic1_keys = list(dic)
    
    workbook = load_workbook(file_path_ex)
    sheet = workbook.active
   
    """sheet.cell(column=1, row=a, value=file_name)
    sheet.cell(column=2, row=a, value=buyer_name)
    sheet.cell(column=3, row=a, value=seller_name)
    sheet.cell(column=4, row=a, value=gst_str)
    sheet.cell(column=5, row=a, value=fssai_str)
    sheet.cell(column=6, row=a, value=str(date))
    sheet.cell(column=7, row=a, value=invoice_total)
    sheet.cell(column=8, row=a, value=str(dic))"""
    for i in range(len(dic1_keys)):
        a = sheet.max_row+1
        sheet.cell(column=1, row=a, value=file_name)
        sheet.cell(column=2, row=a, value=invoice_no)
        sheet.cell(column=3, row=a, value=buyer_name)
        sheet.cell(column=4, row=a, value=seller_name)
        sheet.cell(column=5, row=a, value=gst_str)
        sheet.cell(column=6, row=a, value=fssai_str)
        sheet.cell(column=7, row=a, value=str(date))
        sheet.cell(column=8, row=a, value=invoice_total)
        dic_item = dic.get(dic1_keys[i])
        
        item_obj = dic_item[1].split(":")
        
        item_obj_final = item_obj[-1]
        item_obj_final_strip = item_obj_final.strip()
       
        sheet.cell(column=9, row=a, value=str(item_obj_final_strip))
        hsn_obj = dic_item[2].split(":")
        hsn_obj_final = hsn_obj[-1]
        hsn_obj_final_strip = hsn_obj_final.strip()
        sheet.cell(column=10, row=a, value=str(hsn_obj_final_strip))
      
        weight_obj = dic_item[5].split(":")
        weight_obj_final = weight_obj[-1]
        weight_obj_final_strip = weight_obj_final.strip()
        sheet.cell(column=11, row=a, value=str(weight_obj_final_strip))
        
        rate_obj = dic_item[6].split(":")
        rate_obj_final = rate_obj[-1]
        rate_obj_final_strip = rate_obj_final.strip()
        sheet.cell(column=12, row=a, value=str(rate_obj_final_strip))
       
        amt_obj = dic_item[7].split(":")
        amt_obj_final = amt_obj[-1]
        amt_obj_final_strip = amt_obj_final.strip()
        sheet.cell(column=13, row=a, value=str(amt_obj_final_strip))
        
        if newline_count !=0:
            qty = final_list[i][3]
        else:
            text = dic[1][3]
            pattern = r"BAGS: (\d+)"
            match = re.search(pattern, text)
            qty = match.group(1)



        sheet.cell(column=14, row=a, value=qty)
        workbook.save(file_path_ex)
    
    if (
        sheet.cell(column=1, row=a).value == file_name and
        sheet.cell(column=2, row=a).value == invoice_no and
        sheet.cell(column=3, row=a).value == buyer_name and
        sheet.cell(column=4, row=a).value == seller_name and
        sheet.cell(column=5, row=a).value == gst_str and
        sheet.cell(column=6, row=a).value == fssai_str and
        sheet.cell(column=7, row=a).value == str(date) and
        sheet.cell(column=8, row=a).value == invoice_total and
        sheet.cell(column=9, row=a, value=str(item_obj_final_strip)) and 
        sheet.cell(column=10, row=a, value=str(hsn_obj_final_strip)) and 
        sheet.cell(column=11, row=a, value=str(weight_obj_final_strip)) and
        sheet.cell(column=12, row=a, value=str(rate_obj_final_strip)) and
        sheet.cell(column=13, row=a, value=str(amt_obj_final_strip)) and
        sheet.cell(column=14, row=a, value=qty)
        
    ):
        label2.config(text="Sai-Tanishq Format data uploaded in selected excel sheet.")
        root.after(2500, lambda: label2.config(text=""))
    else:
        label2.config(text="Sai-Tanishq Format data not uploaded")   

def extract_info(format_text):
    global file_path
    if format_text == "Format 1":
        for file in file_path:
            extract_info_from_format_1(file)
    elif format_text == "Format 2":
        for file in file_path:
            extract_info_from_format_2(file)
    elif format_text == "Format 3":
        for file in file_path:
            extract_info_from_format_3(file)

file_path = None 
format_text = None

def open_pdf(selected_format):
    global file_path
    global format_text
    format_text = selected_format 
    file_path = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    if file_path:
#         label.config(text=file_path)
        
        label_text2 = f"Number of files uploaded : {len(file_path)}"
        label.config(text=label_text2)
        label_width2 = len(label_text2) + 2  
        label.config(width=label_width2)

def submit():
    if file_path:
        extract_info(format_text) 
    else:
        label2.config(text="File path is not defined.")

browse_text = StringVar()
button_f1 = Button(root, textvariable=browse_text, command=lambda: open_pdf("Format 1"), bg='light sea green', font="Tahoma", fg="black", height=2, width=20, relief="raised", compound="center")
browse_text.set("Cosoha Format")
button_f1.grid(row=9, column=6, sticky='nw')

browse_text = StringVar()
button_f2 = Button(root, textvariable=browse_text, command=lambda: open_pdf("Format 2"), bg='light sea green', font="Tahoma", fg="black", height=2, width=20, relief="raised", compound="center")
browse_text.set("Rajeshwari Format")
button_f2.grid(row=9, column=6, sticky='n')

browse_text = StringVar()
button_f3 = Button(root, textvariable=browse_text, command=lambda: open_pdf("Format 3"), bg='light sea green', font="Tahoma", fg="black", height=2, width=20, relief="raised", compound="center")
browse_text.set("Sai-Tanishq Format")
button_f3.grid(row=9, column=6, sticky='ne')

button3 = Button(root, text="SUBMIT", font=("Tahoma", 20), command=submit, height=2, width=40, fg='black', bg="DeepSkyBlue3", relief="raised", borderwidth=3, bd=4)
button3.grid(row=10, column=6, columnspan=1, pady=10, padx=10, sticky='n')

button4 = Button(root, fg='black',bg="gray40", text="EXIT", font="Tahoma", command=exit_tk, height=2, width=8, relief="raised", borderwidth=3, bd=1)
button4.grid(row=11, column=6, sticky='nee', pady=10, padx=10)

root.mainloop()